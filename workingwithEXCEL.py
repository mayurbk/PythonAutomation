import openpyxl     ##install openpyxl in pycharm to start working.

#opening excel documents
myexcel = openpyxl.load_workbook("C:\\Users\\mayur\\Downloads\\Automate_the_Boring_Stuff_onlinematerials\\automate_online-materials\\example.xlsx")
print(type(myexcel))

print(myexcel.sheetnames)

#getting cells from sheets
sheet = myexcel['Sheet1']
print(sheet['B1'].value)

#wokring with rows and columns
Value = sheet['B1']
print('Row %s, Column %s is %s' %(Value.row, Value.column,Value.value))   ##try to run from the terminal to check

sheet.cell(row=1, column=2)
for i in range(1,8):
    print(i,sheet.cell(row=i, column=2).value)