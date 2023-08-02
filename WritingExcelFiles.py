import openpyxl

mywb = openpyxl.Workbook()
print(mywb.sheetnames)
sheet = mywb.active
print(sheet.title)
sheet.title = 'Laptop'
print(mywb.sheetnames)
mywb.save('MyExcelFile.xlsx')
mywb.create_sheet()
print(mywb.sheetnames)
mywb.save('MyExcelFile.xlsx')
sheet['A1'] = 'Some value'
print(sheet['A1'].value)
mywb.save('MyExcelFile.xlsx')